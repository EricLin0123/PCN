#!/usr/bin/env python3
"""One-time import of the RA index in main.xlsx into the existing PCN database.

This development migration is never called by the Nuxt application.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PCN_PATTERN = re.compile(r"(?<!\d)(20\d{9})(?!\d)")
PART_SEPARATOR = re.compile(r"[;\r\n]+")
REQUIRED_HEADERS = {
    "RA No.",
    "PCN Number",
    "Related TI Vendor Part Number(s)",
    "Part Count",
    "RA Workbook Filename",
}


class FatalImportError(RuntimeError):
    pass


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def ra_number(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(value)


def declared_count(value: Any) -> int | None:
    try:
        return int(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def parse_parts(value: Any) -> list[tuple[str, str]]:
    unique: dict[str, str] = {}
    for token in PART_SEPARATOR.split(clean(value)):
        display = token.strip().strip('"').strip()
        if display:
            unique.setdefault(display.upper(), display)
    return list(unique.items())


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--database", type=Path, default=ROOT / "data" / "pcn.db")
    parser.add_argument("--workbook", type=Path, default=ROOT / "main.xlsx")
    parser.add_argument("--reset-ra", action="store_true", help="delete only existing RA records before rebuilding them")
    args = parser.parse_args()

    if not args.database.is_file():
        print(f"FATAL: PCN database not found: {args.database}", file=sys.stderr)
        return 1
    if not args.workbook.is_file():
        print(f"FATAL: RA workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    connection = sqlite3.connect(args.database)
    connection.execute("PRAGMA foreign_keys = ON")
    connection.executescript((ROOT / "data" / "schema.sql").read_text(encoding="utf-8"))
    report: Counter = Counter()
    workbook = None
    try:
        existing = connection.execute("SELECT count(*) FROM risk_assessment").fetchone()[0]
        if existing and not args.reset_ra:
            raise FatalImportError(f"database already contains {existing} risk assessments; use --reset-ra only for a development rebuild")
        connection.execute("BEGIN IMMEDIATE")
        if args.reset_ra:
            print("WARNING: --reset-ra deletes all risk assessments and their part links before rebuilding them.", file=sys.stderr)
            connection.execute("DELETE FROM risk_assessment")

        workbook = load_workbook(args.workbook, read_only=True, data_only=True)
        if "RA index" not in workbook.sheetnames:
            raise FatalImportError("main.xlsx does not contain an 'RA index' worksheet")
        worksheet = workbook["RA index"]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [clean(value) for value in next(rows)]
        except StopIteration as exc:
            raise FatalImportError("RA index worksheet is empty") from exc
        missing = sorted(REQUIRED_HEADERS - set(headers))
        if missing:
            raise FatalImportError(f"RA index is missing required headers: {', '.join(missing)}")

        seen_ra: set[str] = set()
        for row_no, values in enumerate(rows, 2):
            if not any(value not in (None, "") for value in values):
                continue
            report["source_rows_read"] += 1
            row = dict(zip(headers, values))
            number = ra_number(row["RA No."])
            if not number:
                raise FatalImportError(f"RA index row {row_no} has no RA number")
            if number in seen_ra:
                raise FatalImportError(f"duplicate RA number {number!r} at row {row_no}")
            seen_ra.add(number)

            raw_pcn = clean(row["PCN Number"])
            match = PCN_PATTERN.search(raw_pcn)
            if not match:
                raise FatalImportError(f"RA {number} has an invalid PCN number: {raw_pcn!r}")
            base = match.group(1)
            pcn = connection.execute("SELECT id FROM pcn WHERE pcn_number_base = ?", (base,)).fetchone()
            if not pcn:
                report["unmatched_pcn_bases"] += 1
                print(f"WARNING: RA {number}: PCN {base} is not in the database", file=sys.stderr)

            parts = parse_parts(row["Related TI Vendor Part Number(s)"])
            expected = declared_count(row["Part Count"])
            if expected is None or expected != len(parts):
                report["part_count_mismatches"] += 1
                print(f"WARNING: RA {number}: Part Count={expected}, parsed={len(parts)}", file=sys.stderr)

            cursor = connection.execute(
                "INSERT INTO risk_assessment(ra_number, pcn_id, pcn_number_base, workbook_filename, source_row) VALUES (?, ?, ?, ?, ?)",
                (number, pcn[0] if pcn else None, base, clean(row["RA Workbook Filename"]), row_no),
            )
            for normalized, _display in parts:
                part = connection.execute("SELECT id FROM ti_part WHERE normalized_part_number = ?", (normalized,)).fetchone()
                if not part:
                    report["parts_not_in_ti_source"] += 1
                    print(f"WARNING: RA {number}: TI part {normalized!r} is not authoritative and was not linked", file=sys.stderr)
                    continue
                if pcn:
                    authoritative = connection.execute(
                        "SELECT 1 FROM pcn_ti_part WHERE pcn_id = ? AND ti_part_id = ?", (pcn[0], part[0])
                    ).fetchone()
                    if not authoritative:
                        report["parts_not_affected_by_pcn"] += 1
                        print(f"WARNING: RA {number}: TI part {normalized!r} is not affected by PCN {base} and was not linked", file=sys.stderr)
                        continue
                connection.execute(
                    "INSERT INTO risk_assessment_ti_part(risk_assessment_id, ti_part_id) VALUES (?, ?)",
                    (cursor.lastrowid, part[0]),
                )

        report.update({
            "risk_assessments_created": connection.execute("SELECT count(*) FROM risk_assessment").fetchone()[0],
            "pcn_bases_with_ra": connection.execute("SELECT count(DISTINCT pcn_id) FROM risk_assessment WHERE pcn_id IS NOT NULL").fetchone()[0],
            "unique_ra_parts": connection.execute("SELECT count(DISTINCT ti_part_id) FROM risk_assessment_ti_part").fetchone()[0],
            "ra_part_links_created": connection.execute("SELECT count(*) FROM risk_assessment_ti_part").fetchone()[0],
        })
        connection.commit()
    except Exception as exc:
        connection.rollback()
        print(f"FATAL: RA import rolled back: {exc}", file=sys.stderr)
        return 1
    finally:
        if workbook:
            workbook.close()
        connection.close()

    fields = ["source_rows_read", "risk_assessments_created", "pcn_bases_with_ra", "unique_ra_parts",
              "ra_part_links_created", "unmatched_pcn_bases", "part_count_mismatches",
              "parts_not_in_ti_source", "parts_not_affected_by_pcn"]
    output = {field: report[field] for field in fields}
    path = args.database.with_name("ra-import-report.json")
    path.write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(output, indent=2))
    print(f"RA import committed. Report written to {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
