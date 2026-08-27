#!/usr/bin/env python3
"""One-time, transactional migration from the two approved source files.

This script is deliberately not used by the Nuxt application. After migration,
data/pcn.db is the sole source of truth.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = ROOT / "data" / "pcn.db"
DEFAULT_TI = ROOT / "source of truth" / "PCN From TI.csv"
DEFAULT_DELTA = ROOT / "source of truth" / "PCN From Delta.xlsx"
PCN_PATTERN = re.compile(r"(?<!\d)(20\d{9})(?!\d)")

TI_HEADERS = {
    "Revision Date Notification",
    "PCN Number",
    "Material",
    "PCN Change Type",
    "PCN Title",
}
DELTA_HEADERS = {
    "APPLY_DATE",
    "FORM_NO",
    "PCN_NO",
    "FORM_STATUS",
    "TOTAL_PNS",
    "AFFECTED_PNS",
    "MAIN_CHANGE_REASON",
    "NOTIFY",
}


class FatalImportError(RuntimeError):
    pass


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().strip('"').strip()


def normalized_part(value: Any) -> str | None:
    result = clean(value).upper()
    return result or None


def normalized_pcn(value: Any) -> str | None:
    match = PCN_PATTERN.search(clean(value))
    return match.group(1) if match else None


def iso_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = clean(value)
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def default_risk(change_type: str) -> str:
    """Conservative initial policy; admins can override each PCN in the GUI."""
    minor_terms = ("INFORMATION", "PACK", "LABEL", "MARK", "DATASHEET", "DATA SHEET")
    upper = change_type.upper()
    return "MINOR" if any(term in upper for term in minor_terms) else "MAJOR"


def int_or_none(value: Any) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_affected_lines(value: Any) -> list[dict[str, Any]]:
    text = clean(value).replace("\r\n", "\n").replace("\r", "\n")
    if not text:
        return []
    parsed = []
    for source_line in text.split("\n"):
        raw = source_line.strip().strip('"').strip()
        if not raw:
            continue
        tokens = raw.split()
        if len(tokens) < 3 or not tokens[0].isdigit():
            parsed.append({"sequence": None, "delta": None, "ti": None, "raw": raw, "status": "UNRESOLVED"})
            continue
        delta = None if tokens[1].lower() == "null" else tokens[1]
        ti = " ".join(tokens[2:]).strip() or None
        parsed.append({"sequence": int(tokens[0]), "delta": delta, "ti": ti, "raw": raw, "status": "PARSED"})
    return parsed


def validate_headers(actual: Iterable[Any], required: set[str], source: str) -> list[str]:
    headers = [clean(value) for value in actual]
    missing = sorted(required - set(headers))
    if missing:
        raise FatalImportError(f"{source} is missing required headers: {', '.join(missing)}")
    return headers


def read_ti(path: Path, report: Counter) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, str]]]:
    pcns: dict[str, dict[str, Any]] = {}
    relationships: dict[str, dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        validate_headers(reader.fieldnames or [], TI_HEADERS, "TI CSV")
        for row_no, row in enumerate(reader, 2):
            report["ti_source_rows_read"] += 1
            base = normalized_pcn(row["PCN Number"])
            if not base:
                report["invalid_pcn_numbers"] += 1
                continue
            change = clean(row["PCN Change Type"]) or "Unspecified"
            existing = pcns.get(base)
            if existing and existing["change_type"] != change:
                report["conflicting_ti_change_types"] += 1
                print(f"WARNING: TI row {row_no}: {base} has conflicting change types "
                      f"{existing['change_type']!r} and {change!r}", file=sys.stderr)
            if not existing:
                pcns[base] = {
                    "date": iso_date(row["Revision Date Notification"]),
                    "title": clean(row["PCN Title"]),
                    "change_type": change,
                }
            part_display = clean(row["Material"])
            part = normalized_part(part_display)
            if part:
                key = f"{base}\0{part}"
                if key in relationships:
                    report["duplicate_records_skipped"] += 1
                else:
                    relationships[key] = {"pcn": base, "normalized": part, "display": part_display}
    return pcns, relationships


def read_delta(path: Path, report: Counter) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = validate_headers(next(rows), DELTA_HEADERS, "Delta workbook")
        except StopIteration as exc:
            raise FatalImportError("Delta workbook is empty") from exc
        forms = []
        seen_forms: set[str] = set()
        for row_no, values in enumerate(rows, 2):
            row = dict(zip(headers, values))
            if not any(value not in (None, "") for value in values):
                continue
            report["delta_source_rows_read"] += 1
            form_no = clean(row["FORM_NO"])
            raw_pcn = clean(row["PCN_NO"])
            base = normalized_pcn(raw_pcn)
            if not base:
                report["invalid_pcn_numbers"] += 1
            if not form_no:
                raise FatalImportError(f"Delta row {row_no} has no FORM_NO")
            if form_no in seen_forms:
                report["duplicate_records_skipped"] += 1
                continue
            seen_forms.add(form_no)
            suffix = None
            if base:
                suffix_text = raw_pcn[raw_pcn.find(base) + len(base):].strip()
                suffix = suffix_text or None
            items = parse_affected_lines(row["AFFECTED_PNS"])
            expected = int_or_none(row["TOTAL_PNS"])
            if expected is not None and expected != len(items):
                report["total_pns_parsing_mismatches"] += 1
                print(f"WARNING: Delta row {row_no} ({form_no}): TOTAL_PNS={expected}, parsed={len(items)}", file=sys.stderr)
            report["unresolved_affected_part_lines"] += sum(item["status"] == "UNRESOLVED" for item in items)
            forms.append({
                "base": base, "raw_pcn": raw_pcn, "suffix": suffix, "form_no": form_no,
                "apply_date": iso_date(row["APPLY_DATE"]), "notify": clean(row["NOTIFY"]) or None,
                "status": clean(row["FORM_STATUS"]) or None,
                "reason": clean(row["MAIN_CHANGE_REASON"]) or None,
                "total": expected, "source_row": row_no, "items": items,
            })
        return forms
    finally:
        workbook.close()


def insert_all(connection: sqlite3.Connection, pcns: dict[str, dict[str, Any]], relationships: dict[str, dict[str, str]], forms: list[dict[str, Any]], report: Counter) -> None:
    for base, item in pcns.items():
        change_name = item["change_type"]
        connection.execute("INSERT OR IGNORE INTO change_type(name, default_risk) VALUES (?, ?)", (change_name, default_risk(change_name)))
        change_id = connection.execute("SELECT id FROM change_type WHERE name = ?", (change_name,)).fetchone()[0]
        connection.execute(
            "INSERT INTO pcn(pcn_number_base, notification_date, title, change_type_id) VALUES (?, ?, ?, ?)",
            (base, item["date"], item["title"], change_id),
        )
    pcn_ids = dict(connection.execute("SELECT pcn_number_base, id FROM pcn"))
    for item in relationships.values():
        connection.execute("INSERT OR IGNORE INTO ti_part(normalized_part_number, display_part_number) VALUES (?, ?)", (item["normalized"], item["display"]))
        part_id = connection.execute("SELECT id FROM ti_part WHERE normalized_part_number = ?", (item["normalized"],)).fetchone()[0]
        connection.execute("INSERT INTO pcn_ti_part(pcn_id, ti_part_id) VALUES (?, ?)", (pcn_ids[item["pcn"]], part_id))
    for form in forms:
        pcn_id = pcn_ids.get(form["base"])
        cursor = connection.execute(
            """INSERT INTO delta_form(pcn_id, delta_pcn_number_base, delta_pcn_number_raw,
               delta_pcn_suffix, form_no, apply_date, notify, form_status, main_change_reason,
               total_pns, source_row) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (pcn_id, form["base"], form["raw_pcn"], form["suffix"], form["form_no"], form["apply_date"],
             form["notify"], form["status"], form["reason"], form["total"], form["source_row"]),
        )
        form_id = cursor.lastrowid
        for item in form["items"]:
            delta_id = None
            delta_normalized = normalized_part(item["delta"])
            if delta_normalized:
                connection.execute("INSERT OR IGNORE INTO delta_part(normalized_part_number, display_part_number) VALUES (?, ?)", (delta_normalized, clean(item["delta"])))
                delta_id = connection.execute("SELECT id FROM delta_part WHERE normalized_part_number = ?", (delta_normalized,)).fetchone()[0]
            connection.execute(
                """INSERT INTO delta_form_item(delta_form_id, sequence_number, delta_part_id,
                   ti_part_number, ti_part_number_normalized, raw_line, parse_status)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (form_id, item["sequence"], delta_id, item["ti"], normalized_part(item["ti"]), item["raw"], item["status"]),
            )

    report.update({
        "unique_pcn_bases": connection.execute("SELECT count(*) FROM pcn").fetchone()[0],
        "unique_ti_parts": connection.execute("SELECT count(*) FROM ti_part").fetchone()[0],
        "unique_delta_parts": connection.execute("SELECT count(*) FROM delta_part").fetchone()[0],
        "pcn_to_ti_part_relationships": connection.execute("SELECT count(*) FROM pcn_ti_part").fetchone()[0],
        "delta_forms_created": connection.execute("SELECT count(*) FROM delta_form").fetchone()[0],
        "delta_form_items_created": connection.execute("SELECT count(*) FROM delta_form_item").fetchone()[0],
        "delta_only_pcn_bases": connection.execute(
            "SELECT count(DISTINCT delta_pcn_number_base) FROM delta_form WHERE delta_pcn_number_base IS NOT NULL AND pcn_id IS NULL"
        ).fetchone()[0],
    })


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--database", type=Path, default=DEFAULT_DB)
    parser.add_argument("--ti", type=Path, default=DEFAULT_TI)
    parser.add_argument("--delta", type=Path, default=DEFAULT_DELTA)
    parser.add_argument("--reset", action="store_true", help="delete the development database before rebuilding it")
    args = parser.parse_args()

    for source in (args.ti, args.delta):
        if not source.is_file():
            print(f"FATAL: required source file not found: {source}", file=sys.stderr)
            return 1
    if args.reset and args.database.exists():
        print(f"WARNING: --reset deletes the development database: {args.database}", file=sys.stderr)
        args.database.unlink()
    args.database.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(args.database)
    report: Counter = Counter()
    try:
        connection.execute("PRAGMA foreign_keys = ON")
        connection.executescript((ROOT / "data" / "schema.sql").read_text(encoding="utf-8"))
        existing = connection.execute("SELECT count(*) FROM pcn").fetchone()[0]
        if existing:
            raise FatalImportError(f"database already contains {existing} PCNs; use --reset only for a development rebuild")
        connection.execute("BEGIN IMMEDIATE")
        pcns, relationships = read_ti(args.ti, report)
        forms = read_delta(args.delta, report)
        if not pcns:
            raise FatalImportError("TI source produced no valid PCNs")
        insert_all(connection, pcns, relationships, forms, report)
        connection.commit()
    except Exception as exc:
        connection.rollback()
        print(f"FATAL: import rolled back: {exc}", file=sys.stderr)
        return 1
    finally:
        connection.close()

    fields = [
        "ti_source_rows_read", "delta_source_rows_read", "unique_pcn_bases", "unique_ti_parts",
        "unique_delta_parts", "pcn_to_ti_part_relationships", "delta_forms_created",
        "delta_form_items_created", "delta_only_pcn_bases", "invalid_pcn_numbers",
        "conflicting_ti_change_types", "total_pns_parsing_mismatches",
        "unresolved_affected_part_lines", "duplicate_records_skipped",
    ]
    output = {field: report[field] for field in fields}
    report_path = args.database.with_name("import-report.json")
    report_path.write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(output, indent=2))
    print(f"Import committed. Report written to {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
